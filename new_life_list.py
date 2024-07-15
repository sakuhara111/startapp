import datetime
import glob
import os

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

MAXROW=30-1
MAXCOL=7
USECOLS=["種類","まさ","かずは","買うかどうか","優先度(買う時期)","メモ","参考URL"]
TABLIST=['家電','家具・インテリア', 'キッチン', '日用品（キッチン以外）']
# ataFrameが一度だけロードされ、その結果がキャッシュに格納される
@st.cache_data
def read_df(st_name,excel_file):
    # 特定のシートを読み込む
    # Excelファイルを読み込む
    # 最初の5行だけを読み込む
    # 特定の列のみを読み込む,usecols=['Column1', 'Column2']
    df = pd.read_excel(
        excel_file, sheet_name=st_name, 
        nrows=MAXROW
        )

    # データフレームの内容を表示
    print(df)
    return df

def set_tab(tab_name,num,excel_file):
    # def df_callback():
    #     st.session_state["count"] += 1

    # if "count" not in st.session_state:
    #     st.session_state["count"] = 0
    st.write(tab_name)
    df=read_df(tab_name,excel_file)
    df=df.filter(items=USECOLS)
    df_select=read_df('選択リスト',excel_file)
    df_select=df_select.fillna('')
    select_list1=df_select["所持してるか"].tolist()
    # select_list1= filter(lambda a: a != '', select_list1)
    select_list2=df_select["買うかどうか"].tolist()
    # select_list2= filter(lambda a: a != '', select_list2)
    select_list3=df_select["優先度(買う時期)"].tolist()
    # select_list3= filter(lambda a: a != '', select_list3)
    
    # セッションステートにデータフレームを保存
    if f"df{num}" not in st.session_state:
        st.session_state[f"df{num}"] = df
    
    # 編集内容を保存する関数
    def save_edits(num):
        st.session_state[f"df{num}"] = edited_df
        # データをCSVとして保存
        csv_filename = f"data{num}.csv"
        st.session_state[f"df{num}"].to_csv(csv_filename, index=False)
        st.success("データが保存されました")
        st.write(st.session_state[f"df{num}"])
        
    
    # , '家具・インテリア', 'キッチン', '日用品（キッチン以外）', '選択リスト')
    # データフレームの内容を表示
    # データフレームの編集
     # データエディタでデータフレームを表示・編集
    edited_df = st.data_editor(
        st.session_state[f"df{num}"]
        ,column_config={
            "種類": st.column_config.Column(
                "種類",
                help="物を入力",
                width="medium",
                # default="何かあれば記入してね.",
                required=False,
            ),
            "まさ": st.column_config.SelectboxColumn(
                "まさ",
                help="選択",
                options=select_list1,
                required=False,
            ),
            "かずは": st.column_config.SelectboxColumn(
                "かずは",
                help="選択",
                options=select_list1,
                required=False,
            ),
            "買うかどうか": st.column_config.SelectboxColumn(
                "買うかどうか",
                help="選択",
                options=select_list2,
                required=False,
            ),
            "優先度(買う時期)": st.column_config.SelectboxColumn(
                "優先度(買う時期)",
                help="選択",
                options=select_list3,
                required=False,
            ),
            "メモ": st.column_config.Column(
                "メモ",
                help="入力",
                width="medium",
                # default="何かあれば記入してね.",
                required=False,
            ),
            "URL":st.column_config.LinkColumn(
                "URL",
                help="URL入力", 
                required=False,
                display_text="Open URL"
                )
        },
        use_container_width=True,
        hide_index=False,
        num_rows="dynamic",
        # on_change=df_callback,
        key="df2"+str(num)
    )
    st.write(num)
    st.write('以下補足')
    st.write('所持してるか:有り/無し/買換？')
    st.write('買うかどうか:買う/買わない/保留')
    st.write('優先度(買う時期):引越直後/1か月以内/徐々に/')
    # 更新されたデータフレームの表示
    st.write('入力が完了し、他のタブへ移動する場合は保存ボタンをクリック')
    # データフレームの編集
    st.button('変更を保存', on_click=lambda :save_edits(num),key="button"+str(num))

    # st.dataframe(st.session_state.original_data)
 

def add_csv_to_excel(excel_file, csv_files):
    """
    既存のExcelファイルにCSVファイルのデータを追加する関数
    
    Parameters:
    - excel_file: str, Excelファイルのパス
    - csv_files: list, 追加するCSVファイルのパスのリスト
    - tablist: list, 各CSVファイルに対応するシート名のリスト
    """
    # 既存のExcelファイルを読み込む
    wb = load_workbook(excel_file)
    
    # 各CSVファイルを読み込み、Excelファイルに追加
    for i, csv_file in enumerate(csv_files):
        # CSVファイルを読み込み、ヘッダーを除く
        df = pd.read_csv(csv_file).iloc[1:]
        # シート名を設定
        sheet_name = tablist[i]
        
        # シートが存在しない場合は新規作成
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        
        # シートを選択
        sheet = wb[sheet_name]
        
        # データフレームの内容をシートに書き込む
        for r_idx, row in enumerate(df.values, start=2):  # 2行目から開始
            for c_idx, value in enumerate(row, start=1):  # 1列目から開始
                sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # 変更を保存
    wb.save(excel_file)


def update_data(base_dir,excel_file):
    # CSVファイルのパスを取得
    csv_pattern = os.path.join(base_dir, 'data*.csv')
    csv_files = glob.glob(csv_pattern)
    if len(csv_files)==4:
        add_csv_to_excel(excel_file, csv_files)        
        return len(csv_files)
    else:
        return len(csv_files)
def main():
    # 実行ファイルのディレクトリを取得
    base_dir = os.path.dirname(os.path.abspath(__file__))
    # 既存のExcelファイルのパスを取得
    excel_file = os.path.join(base_dir, "引越しと生活のリスト.xlsx")
    # CSVファイルのパスを取得
    csv_pattern = os.path.join(base_dir, 'data*.csv')
    csv_files = glob.glob(csv_pattern)
    # 各CSVファイルを読み込み、Excelファイルに追加
    for csv_file in csv_files:
        os.remove(csv_file)
    
        
    # ページ情報、基本的なレイアウト
    st.set_page_config(
        page_title="引越し・生活用品リスト",
        page_icon="🧊",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    # １ページ目表示
    # st.sidebar.title("test_streamlit")
    # st.markdown("##完了したら保存ボタンを押す")
    # st.sidebar.button("保存", on_click=change_page)

    # サイドバー
    # select_option = st.sidebar.selectbox(
    #     "セレクトボックス", ("Email", "Home phone", "Mobile phone")
    # )
    # タイトル。最もサイズが大きい。ページタイトル向け
    # st.title('引越し・生活用品リスト')

    # ヘッダ。２番目に大きい。項目名向け
    st.header('引越し・生活用品リスト')

    # サブレベルヘッダ。３番目に大きい。小項目向け
    st.subheader('下のタブで項目の切替ができるから適当に入れてね')

    st.write('4つのタブの入力が完了したら完了ボタンをクリック')
    if st.button("完了",key="end"):
        csv_files_num=update_data(base_dir,excel_file)
        if csv_files_num==4:
            st.write(csv_files_num)
            st.success('更新されました、このアプリを閉じて下さい')
        else:
            st.warning('Please input a name.:'+str(csv_files_num))

    else:
        st.write("データ更新前")
    
    # 普通のテキスト。Html や Markdown のパースはしない。
    # st.text('Text')
    # タブ
    tabs = st.tabs(TABLIST)
    for i, tab_name in enumerate(TABLIST):
        with tabs[i]:
            set_tab(tab_name, i+1, excel_file)
    
if __name__ == '__main__':
    main()
    
